import csv
import os
import platform
from io import BytesIO, StringIO
from datetime import datetime

from flask import Flask, abort, make_response, request, render_template
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Side, Border
from openpyxl.utils import get_column_letter


app = Flask(__name__)

@app.route("/seplan_or_download/", methods=["GET"])
@app.route("/seplan_or_download", methods=["GET"])
def seplan_or_download():
    # Obter parâmetros de filtro da URL
    try:
        # --- NOVO PRINT DE DEBUG AQUI ---
        gnd_raw_from_url = request.args.get("gnd")
        print(f"\n--- DEBUG INÍCIO: Valor RAW de 'gnd' da URL: '{gnd_raw_from_url}' (Tipo: {type(gnd_raw_from_url)}) ---")
        # --------------------------------

        excel_file_path = os.path.join(app.root_path, 'templates', 'DadosOrcamentoConsolidadoGeral7.xlsx')
        
        if not os.path.exists(excel_file_path):
            abort(404, description=f"Arquivo de base de dados '{os.path.basename(excel_file_path)}' não encontrado na pasta 'templates'.")

        wb = load_workbook(excel_file_path)
        ws = wb.active

        # --- 1. Modificação para a Célula B4 (Data de Referência) ---
        today_date = datetime.now().strftime("%d/%m/%Y")
        ws['B4'] = f"Data de referência: {today_date}"

        # --- Definir a Estrutura do Excel para a Lógica de Filtragem e Formatação ---
        header_row_index = 9  # Assumindo que os cabeçalhos da tabela estão na linha 9
        data_start_row = 10   # Assumindo que os dados começam na linha 10
        data_start_col = 2    # Assumindo que os dados e cabeçalhos começam na Coluna B (índice 2)

        excel_header_map = {}
        max_excel_col = ws.max_column
        for col_idx in range(data_start_col, max_excel_col + 1):
            header_cell_value = ws.cell(row=header_row_index, column=col_idx).value
            if header_cell_value is not None:
                excel_header_map[str(header_cell_value)] = col_idx

        # Mapeamento dos parâmetros da URL para os nomes EXATOS das colunas no Excel
        FILTER_MAP = {
            "ano": "Ano", 
            "mes": "Mês", 
            "descricaoacao": "Ação e Subtítulo", 
            "descricaoorcamentaria": "Descrição Orçamentária", 
            "gnd": "GND",
            "programaticaorcamentaria": "Programática (Programa, Ação e Subtítulo)", 
            "descricaoprograma": "Programa", 
            "codigo": "Código", 
            "codigo_fonte": "Código Fonte", 
            "descricao_fonte": "Descrição Fonte",
            "funcao_subfuncao": "Função e Subfunção",
            "esfera": "Esfera",
            "dotacao_inicial": "Dotação Inicial",
            "acrescimos": "Acréscimos",
            "decrescimos": "Decréscimos",
            "dotacao_atualizada": "Dotação Atualizada",
            "contingenciado": "Contingenciado",
            "provisao": "Provisão",
            "destaque": "Destaque",
            "dotacao_liquida": "Dotação Líquida",
            "empenhado": "Empenhado",
            "empenhado_porcento": "% Empenhado",
            "liquidado": "Liquidado",
            "liquidado_porcento": "% Liquidado",
            "pago": "Pago",
            "pago_porcento": "% Pago"
        }
        
        # --- Prepara os parâmetros de filtro de forma mais robusta ---
        filter_params = {}
        for param_name, _ in FILTER_MAP.items():
            param_value_raw = request.args.get(param_name)
            
            if param_value_raw is not None:
                cleaned_value = param_value_raw.strip()
                if cleaned_value not in ["", "%"]:
                    filter_params[param_name] = cleaned_value.lower() 
                else:
                    filter_params[param_name] = None
            else:
                filter_params[param_name] = None
        
        print(f"\n--- DEBUG: Parâmetros de Filtro Ativos ---")
        for k, v in filter_params.items():
            if v is not None:
                print(f"  {k}: '{v}'")
        print(f"-----------------------------------------\n")


        # Definir estilos para aplicação de bordas e fonte nas células de dados
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))
        data_font = Font(name="Arial", size=8, color="000000") # Fonte padrão para os dados

        for r_idx in range(data_start_row, ws.max_row + 1):
            row_hidden = False
            
            # Itera sobre o FILTER_MAP para garantir a ordem e quais colunas verificar
            for param_name, excel_header_name in FILTER_MAP.items():
                param_value = filter_params.get(param_name)

                # Se o parâmetro não foi fornecido (é None), pula este filtro para a linha atual
                if param_value is None:
                    continue 

                col_idx = excel_header_map.get(excel_header_name)
                # Se o nome do cabeçalho do Excel não for encontrado, pula este filtro
                if col_idx is None:
                    continue 

                excel_cell_value = ws.cell(row=r_idx, column=col_idx).value
                
                is_match = False
                
                # --- DEBUG: Valores sendo comparados ---
                print(f"  Linha {r_idx}, Coluna '{excel_header_name}' (param: '{param_name}')")
                print(f"    Valor do Parâmetro (URL): '{param_value}' (Tipo: {type(param_value)})")
                print(f"    Valor da Célula Excel: '{excel_cell_value}' (Tipo: {type(excel_cell_value)})")


                # --- Lógica de filtro para Ano e Mês (com prioridade numérica) ---
                if param_name in ["ano", "mes"]:
                    try:
                        # Tenta converter ambos para inteiro para comparação numérica
                        excel_val_int = int(excel_cell_value) if excel_cell_value is not None else None
                        param_val_int = int(param_value) 

                        if excel_val_int is not None and param_val_int == excel_val_int:
                            is_match = True
                            print(f"    --> MATCH! (Numérico: {param_val_int} == {excel_val_int})")
                        else:
                             print(f"    --> NÃO MATCH (Numérico: {param_val_int} != {excel_val_int})")
                    except (ValueError, TypeError):
                        # Se a conversão para inteiro falhar, tenta comparação de string exata
                        cell_val_str = str(excel_cell_value).strip().lower() if excel_cell_value is not None else ""
                        if param_value == cell_val_str:
                            is_match = True
                            print(f"    --> MATCH! (String Exata: '{param_value}' == '{cell_val_str}')")
                        else:
                            print(f"    --> NÃO MATCH (String Exata: '{param_value}' != '{cell_val_str}')")

                else: # Para outros campos de texto: lógica de "contém"
                    cell_val_str = str(excel_cell_value).strip().lower() if excel_cell_value is not None else ""
                    if param_value in cell_val_str:
                        is_match = True
                        print(f"    --> MATCH! (Contém: '{param_value}' IN '{cell_val_str}')")
                    else:
                        print(f"    --> NÃO MATCH (Contém: '{param_value}' NOT IN '{cell_val_str}')")
                
                # Se este filtro ATIVO NÃO corresponder, esconde a linha e para de verificar esta linha
                if not is_match:
                    row_hidden = True
                    print(f"    --- Linha {r_idx} SERÁ ESCONDIDA devido a '{excel_header_name}' ---")
                    break 

            ws.row_dimensions[r_idx].hidden = row_hidden
            if not row_hidden:
                print(f"  *** Linha {r_idx} PERMANECERÁ VISÍVEL ***")
            print("-" * 50) # Separador para facilitar a leitura do debug
            
            # --- APLICAR FORMATAÇÃO BÁSICA (Bordas e Fonte) para Linhas VISÍVEIS ---
            if not row_hidden:
                for c_idx in range(data_start_col, max_excel_col + 1): # Itera por todas as colunas de dados
                    cell = ws.cell(row=r_idx, column=c_idx)
                    cell.border = thin_border # Aplica as bordas
                    cell.font = data_font    # Aplica a fonte

                    if isinstance(cell.value, (int, float)):
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")


        # --- Aplicar Filtros (Triângulos) na Tabela de Dados ---
        filter_start_col_letter = get_column_letter(data_start_col)
        filter_start_row = header_row_index

        filter_end_col_letter = get_column_letter(max_excel_col) 
        filter_end_row = ws.max_row

        filter_range = f"{filter_start_col_letter}{filter_start_row}:{filter_end_col_letter}{filter_end_row}"
        ws.auto_filter.ref = filter_range

        # --- Congelar Painéis ---
        ws.freeze_panes = f"{get_column_letter(data_start_col)}{header_row_index + 1}"


        xlsx_data = BytesIO()
        wb.save(xlsx_data)
        xlsx_data.seek(0)

        response = make_response(xlsx_data.getvalue())
        response.headers["Content-Disposition"] = "attachment; filename=DadosOrcamentoConsolidado.xlsx"
        response.headers["Content-type"] = (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        return response

    except Exception as e:
        abort(500, description=f"Erro ao processar o arquivo Excel: {str(e)}")

    finally:
        pass


@app.route("/seplan_or/", methods=["GET"])
@app.route("/seplan_or", methods=["GET"])
def seplan_or():
    return render_template("seplan_or.html")


@app.route("/", methods=["GET"])
def index():
    return "Extrator DGT - Versão Pública - Gerador de Relatórios Orçamentários."
